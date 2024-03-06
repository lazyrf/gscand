import datetime
import logging
import sys
import argparse
import time
import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.utils import get_column_letter
from itertools import groupby
from farmer_db import easyfarmer

DEVEL = True
gid = '2037314b-50525007-440020'

def validate_date(date_str):
    try:
        return datetime.datetime.strptime(date_str, "%Y-%m-%d")
    except ValueError:
        msg = f"not a valid date: {date_str}"
        raise argparse.ArgumentTypeError(msg)


def logger_init():
    logger = logging.getLogger()
    logger.handlers = []
    logger.setLevel(logging.DEBUG)

    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')

    stdout_handler = logging.StreamHandler(sys.stdout)
    stdout_handler.setLevel(logging.DEBUG)
    stdout_handler.setFormatter(formatter)
    logger.addHandler(stdout_handler)

    file_handler = logging.FileHandler(args.output if args.output else "gscand.log" if DEVEL else "/var/log/gscand/gscand.log")
    file_handler.setLevel(logging.DEBUG)
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)


def ng_value_fn(x):
    return None if x == -9999 else x


def check_sensor(g, snr_type, target_date, sheet, c):
    start_dt = datetime.datetime.combine(target_date, datetime.datetime.min.time())
    end_dt = datetime.datetime.combine(target_date, datetime.datetime.max.time())

    print(f"[{snr_type}] Checking data from {start_dt} to {end_dt}")

    i = 2
    dns = ez.get_dns(g, snr_type)
    for dn in dns:
        ds = ez.get_data(g.gateway_id, dn.dnid, int(start_dt.timestamp()), int(end_dt.timestamp()))

        sheet.cell(row=i, column=1).value = dn.name
        if ds is not None:
            ds_dezip = list(zip(*ds))
            ds_values = list(ds_dezip[0])

            voltage = 0
            wm_diff = 0

            filter_values = filter(ng_value_fn, ds_values)
            valid_values = [v for v in filter_values]

            if dn.dcid == 'dc0041':
                if len(valid_values) > 0:
                    voltage = valid_values[-1]
            elif dn.dcid == 'dc0051':
                if len(valid_values) >= 2:
                    wm_diff = float(valid_values[-1]) - float(valid_values[0])

            values_count = len(ds_values)
            failed_count = ds_values.count(-9999)

            is_abnormal = False
            clean_values = [0 if v != -9999 else v for v in ds_values]
            group_values = [(i, len(list(j))) for i, j in groupby(clean_values)]
            for el_value, el_count in group_values:
                if el_value == -9999.0 and el_count > 24:
                    is_abnormal = True
                    break

            if values_count - failed_count == 0:
                sheet.cell(row=i, column=c).value = 'X'
                sheet.cell(row=i, column=c).fill = redFill
            elif failed_count > 0 and is_abnormal:
                sheet.cell(row=i, column=c).value = '?' if dn.dcid != 'dc0041' else f'{voltage}'
                sheet.cell(row=i, column=c).fill = greenFill
            else:
                if dn.dcid == 'dc0051':
                    sheet.cell(row=i, column=c).value = 0 if wm_diff == 0.0 else f'{round(wm_diff,2)}'
                else:
                    sheet.cell(row=i, column=c).value = f'OK' if dn.dcid != 'dc0041' else f'{voltage}'
        i = i + 1


def main_task(g, target_date, c):
    sheet0.column_dimensions[get_column_letter(c)].width = 10
    sheet1.column_dimensions[get_column_letter(c)].width = 10
    sheet2.column_dimensions[get_column_letter(c)].width = 10

    sheet0.cell(row=1, column=c).value = f'{target_date.month}/{target_date.day}'
    sheet1.cell(row=1, column=c).value = f'{target_date.month}/{target_date.day}'
    sheet2.cell(row=1, column=c).value = f'{target_date.month}/{target_date.day}'

    check_sensor(g, easyfarmer.EasyFarmer.SNR_TYPE_ESP, target_date, sheet0, c)
    if mode != 1:
        check_sensor(g, easyfarmer.EasyFarmer.SNR_TYPE_LORA, target_date, sheet1, c)
    check_sensor(g, easyfarmer.EasyFarmer.SNR_TYPE_WM, target_date, sheet2, c)



if __name__ == '__main__':
    ez = easyfarmer.EasyFarmer()
    g = ez.get_gw(gid)

    today = datetime.date.today()
    yesterday = today - datetime.timedelta(days=1)
    delta = None

    parser = argparse.ArgumentParser(description="gscan daemon")
    parser.add_argument("-o", "--output", help="Output log file")
    parser.add_argument("-d", "--date", help="Date", type=validate_date)
    parser.add_argument("-s", "--start", help="Start datetime", type=validate_date)
    parser.add_argument("-e", "--end", help="End datetime", type=validate_date)
    parser.add_argument("-m", "--mode", help="Mode")
    args = parser.parse_args()

    if args.start and args.end is None:
         sys.exit("End date miss. Terminate gscand daemon.")
    elif args.end and args.start is None:
         sys.exit("Start date miss. Terminate gscand daemon.")
    elif args.start and args.end and args.date:
         sys.exit("Select date range or specified date. Terminate gscand daemon.")

    if args.start and args.end:
        start_dt = datetime.datetime.combine(args.start, datetime.datetime.min.time())
        end_dt = datetime.datetime.combine(args.end, datetime.datetime.max.time())
        delta = end_dt - start_dt
    elif args.date:
        start_dt = datetime.datetime.combine(args.date, datetime.datetime.min.time())
        end_dt = datetime.datetime.combine(args.date, datetime.datetime.max.time())
    else:
        start_dt = datetime.datetime.combine(yesterday, datetime.datetime.min.time())
        end_dt = datetime.datetime.combine(yesterday, datetime.datetime.max.time())

    mode = int(args.mode) if args.mode else 0

    logger_init()
    logging.info("-" * 50)
    logging.info("<<< gscan daemon >>>")
    logging.info(f"Mode: {mode}")
    logging.info(f"{start_dt} - {end_dt}")
    logging.info("-" * 50)

    logging.info("gscand starting...")

    now_t = int(time.time())

    workbook = openpyxl.Workbook()


    redFill = PatternFill(start_color='FFFF6B6B',
                   end_color='FFFF6B6B',
                   fill_type='solid')
    greenFill = PatternFill(start_color='FF63e6be',
                   end_color='FF63e6be',
                   fill_type='solid')

    sheet0 = workbook.create_sheet('氣象站', 0)
    sheet0.cell(row=1, column=1).value = '感測器'
    sheet0.column_dimensions['A'].width = 20

    sheet1 = workbook.create_sheet('水位計', 1)
    sheet1.cell(row=1, column=1).value = '感測器'
    sheet1.column_dimensions['A'].width = 20

    sheet2 = workbook.create_sheet('水錶', 2)
    sheet2.cell(row=1, column=1).value = '感測器'
    sheet2.column_dimensions['A'].width = 20

    c = 2

    if delta is not None:
        for i in range(delta.days + 1):
            target_date = start_dt.date() + datetime.timedelta(days=i)
            main_task(g, target_date, c)
            c = c + 1
    else:
        main_task(g, start_dt.date(), c)

    workbook.save("test.xlsx")

    logging.info("gscand terminated.")
